import calendar
import re
import unicodedata
from collections import defaultdict
from datetime import datetime, date
from pathlib import Path
from typing import Optional, List, Tuple
import tkinter as tk
from tkinter import filedialog, messagebox

try:
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font
except ImportError:
    load_workbook = None
    PatternFill = None
    Font = None

try:
    from PIL import Image
except ImportError:
    Image = None


def clean_name(text: str) -> str:
    """Loại bỏ số điện thoại khỏi tên, giữ nguyên dấu và format"""
    if not text:
        return ""
    
    text = str(text)
    
    # Loại bỏ số điện thoại dạng (0xxx xxx xxx) hoặc (0xxx-xxx-xxx)
    text = re.sub(r'\s*\([0-9\s\-\.]+\)\s*', '', text)
    
    return re.sub(r"\s+", " ", text).strip()


def normalize_text(text: str) -> str:
    """Chuẩn hóa text: bỏ dấu, uppercase, collapse spaces, loại bỏ số điện thoại"""
    if not text:
        return ""
    
    text = clean_name(text)
    
    normalized = unicodedata.normalize("NFD", text)
    no_accent = "".join(ch for ch in normalized if unicodedata.category(ch) != "Mn")
    return re.sub(r"\s+", " ", no_accent.upper()).strip()


def parse_month_from_filename(filepath: Path) -> Optional[Tuple[int, int]]:
    """
    Parse tháng/năm từ tên file: LenhDieuDongT11-2025 → (11, 2025)
    """
    name = filepath.stem
    match = re.search(r"T(\d{1,2}).*?(\d{4})", name, re.IGNORECASE)
    if match:
        month = int(match.group(1))
        year = int(match.group(2))
        if 1 <= month <= 12 and 2000 <= year <= 2099:
            return month, year
    return None


def parse_date_range(date_str, target_month: int, target_year: int) -> List[int]:
    """
    Parse chuỗi ngày thành list các ngày trong tháng.
    VD: 
    - "03-06/11/2025" (tháng 11, liên tục) → [3, 4, 5, 6]
    - "10-11-13/12/2025" (tháng 12, không liên tục) → [10, 11, 13]
    - "01/11/2025" (tháng 11) → [1]
    - datetime(2025, 11, 5) (tháng 11) → [5]
    """
    
    # Nếu là datetime/date object
    if isinstance(date_str, datetime):
        date_str = date_str.date()
    if isinstance(date_str, date):
        if date_str.month == target_month and date_str.year == target_year:
            return [date_str.day]
        return []
    
    if not date_str:
        return []
    
    text = str(date_str).strip()

    # Pattern: "03-06/11/2025" hoặc "10-11-13/12/2025"
    pattern = r"([\d\-]+)/(\d{1,2})/(\d{4})"
    match = re.search(pattern, text)
    if match:
        days_part = match.group(1)
        month = int(match.group(2))
        year = int(match.group(3))
        
        if month == target_month and year == target_year:
            # Filter out empty strings from split
            day_numbers = [int(x) for x in days_part.split('-') if x.strip()]
            
            if not day_numbers:
                return []
            
            # Nếu chỉ có 2 số: coi là range liên tục (VD: "03-06" → [3,4,5,6])
            # Nếu >= 3 số: coi là các ngày riêng lẻ (VD: "10-11-13" → [10,11,13])
            if len(day_numbers) == 2:
                return list(range(min(day_numbers), max(day_numbers) + 1))
            else:
                return sorted(day_numbers)
        return []
    
    # Pattern: "01/11/2025"
    for fmt in ["%d/%m/%Y", "%d-%m-%Y"]:
        try:
            parsed = datetime.strptime(text, fmt).date()
            if parsed.month == target_month and parsed.year == target_year:
                return [parsed.day]
        except ValueError:
            continue
    
    return []


def parse_location(location_str: str) -> str:
    """
    Parse địa điểm công tác - GIỮ NGUYÊN text gốc.
    Input: "Cần Thơ (ST+HG) + An Giang (KG)"
    Output: "Cần Thơ (ST+HG) + An Giang (KG)"
    
    Input: "TP HCM"
    Output: "TP HCM"
    """
    if not location_str:
        return ""
    
    return str(location_str).strip()


def collect_work_data(input_path: Path, target_month: int, target_year: int):
    """
    Đọc sheet "Nhap Lieu", trả về dict cho cả tài xế (cột R) và công nhân (cột S):
    {
        "person_name": {
            "days": {1, 3, 5, ...},  # set of days
            "notes": ["1(TP HCM)", "3-7(Cần Thơ (ST+HG) + An Giang (KG))", ...]
        }
    }
    """
    wb = load_workbook(input_path, data_only=True, read_only=True)
    
    # Find sheet "Nhap Lieu"
    if "Nhap Lieu" in wb.sheetnames:
        ws = wb["Nhap Lieu"]
    else:
        ws = wb.active
    
    # Find header row (row containing "TAI XE" or similar)
    header_row = 1
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10), 1):
        for cell in row:
            if cell.value and "TAI XE" in normalize_text(str(cell.value)):
                header_row = row_idx
                break
        if header_row == row_idx:
            break
    
    # Find columns R, S, U, W (or by header name)
    driver_col = None
    worker_col = None
    location_col = None
    time_col = None
    
    header_cells = list(ws[header_row])
    for cell in header_cells:
        norm = normalize_text(str(cell.value)) if cell.value else ""
        if "TAI XE" in norm and not driver_col:
            driver_col = cell.column
        elif "CONG NHAN" in norm and not worker_col:
            worker_col = cell.column
        elif "DIA DIEM" in norm and "CONG TAC" in norm and not location_col:
            location_col = cell.column
        elif "THOI GIAN" in norm and "CONG TAC" in norm and not time_col:
            time_col = cell.column
    
    # Fallback to R, S, U, W
    if not driver_col:
        driver_col = 18  # R
    if not worker_col:
        worker_col = 19  # S
    if not location_col:
        location_col = 21  # U
    if not time_col:
        time_col = 23  # W
    
    result = defaultdict(lambda: {"days": set(), "notes": [], "original_name": ""})
    
    for row in ws.iter_rows(min_row=header_row + 1):
        driver = row[driver_col - 1].value
        worker = row[worker_col - 1].value
        location = row[location_col - 1].value
        time_value = row[time_col - 1].value
        
        if not time_value:
            continue
        
        days = parse_date_range(time_value, target_month, target_year)
        
        if not days:
            continue
        
        # Build note: "1(TP HCM)" hoặc "3-7(Cần Thơ)" hoặc "10+11+13(An Giang)"
        location_text = parse_location(str(location) if location else "")
        if len(days) == 1:
            note = f"{days[0]}({location_text})"
        else:
            # Kiểm tra ngày có liên tục không
            sorted_days = sorted(days)
            is_consecutive = all(sorted_days[i] + 1 == sorted_days[i + 1] for i in range(len(sorted_days) - 1))
            
            if is_consecutive:
                # Liên tục: "3-7"
                note = f"{min(sorted_days)}-{max(sorted_days)}({location_text})"
            else:
                # Không liên tục: "10+11+13"
                note = f"{'+'.join(map(str, sorted_days))}({location_text})"
        
        # Process driver (cột R)
        if driver:
            driver_key = normalize_text(str(driver))
            if not result[driver_key]["original_name"]:
                result[driver_key]["original_name"] = clean_name(str(driver))
            result[driver_key]["days"].update(days)
            if note not in result[driver_key]["notes"]:
                result[driver_key]["notes"].append(note)
        
        # Process worker (cột S)
        if worker:
            worker_key = normalize_text(str(worker))
            if not result[worker_key]["original_name"]:
                result[worker_key]["original_name"] = clean_name(str(worker))
            result[worker_key]["days"].update(days)
            if note not in result[worker_key]["notes"]:
                result[worker_key]["notes"].append(note)
    
    return dict(result)


def find_template_file(base_dir: Path, month: int, year: int) -> Optional[Path]:
    """
    Tìm file template "Cham cong {MM} {YYYY}.xlsx"
    VD: "Cham cong 11 2025.xlsx"
    """
    candidates = [
        base_dir / f"Cham cong {month:02d} {year}.xlsx",
        base_dir / f"Cham cong {month} {year}.xlsx",
        base_dir / "template.xlsx",
    ]
    
    for path in candidates:
        if path.exists():
            return path
    
    # Tìm file bất kỳ bắt đầu "Cham cong"
    for path in base_dir.glob("Cham cong*.xlsx"):
        if "_ket_qua" not in path.name.lower():
            return path
    
    return None


def get_sundays_in_month(year: int, month: int) -> List[int]:
    """
    Trả về list các ngày Chủ Nhật trong tháng.
    VD: get_sundays_in_month(2025, 11) → [2, 9, 16, 23, 30]
    """
    days_in_month = calendar.monthrange(year, month)[1]
    sundays = []
    for day in range(1, days_in_month + 1):
        if calendar.weekday(year, month, day) == 6:  # 6 = Sunday
            sundays.append(day)
    return sundays


def fill_sunday_sheet(wb, work_data: dict, target_month: int, target_year: int):
    """
    Fill dữ liệu vào sheet "Chủ Nhật" cho tài xế/công nhân làm việc ngày Chủ Nhật.
    Tự động điền header các ngày CN và danh sách người vào sheet.
    """
    # Tìm sheet "Chủ Nhật"
    sunday_sheet = None
    for sheet_name in wb.sheetnames:
        norm_name = normalize_text(sheet_name)
        if "CHU NHAT" in norm_name or "CN" == norm_name:
            sunday_sheet = wb[sheet_name]
            break
    
    if not sunday_sheet:
        return
    
    # Lấy danh sách các ngày Chủ Nhật trong tháng
    sundays = get_sundays_in_month(target_year, target_month)
    if not sundays:
        return
    
    # Tìm hàng header (cố định là row 5 theo template)
    header_row = 5
    
    # Điền các ngày Chủ Nhật vào header (bắt đầu từ cột D)
    sunday_start_col = 4  # Column D
    sunday_cols = {}  # {day: col_idx}
    for idx, day in enumerate(sundays):
        col_idx = sunday_start_col + idx
        cell = sunday_sheet.cell(header_row, col_idx)
        cell.value = f"{day:02d}/{target_month:02d}/{target_year}"
        sunday_cols[day] = col_idx
    
    # Tìm cột "Ghi chú" (hoặc dùng cột I như trong template)
    note_col = 9  # Column I
    
    # Lọc ra những người làm việc ngày Chủ Nhật
    sunday_workers = {}
    for person_key, work_info in work_data.items():
        sunday_days = [day for day in sundays if day in work_info["days"]]
        if sunday_days:
            sunday_workers[person_key] = {
                "days": sunday_days,
                "notes": work_info["notes"]
            }
    
    if not sunday_workers:
        return
    
    # Điền dữ liệu vào các rows (bắt đầu từ row 6)
    data_start_row = 6
    row_idx = data_start_row
    
    for person_key, work_info in sorted(sunday_workers.items()):
        # Cột A: Họ và tên (dùng tên gốc)
        original_name = work_data[person_key].get("original_name", person_key)
        sunday_sheet.cell(row_idx, 1).value = original_name
        
        # Cột C: Số ngày
        sunday_sheet.cell(row_idx, 3).value = len(work_info["days"])
        
        # Điền 'X' vào các cột ngày Chủ Nhật
        sunday_notes = []
        for day in work_info["days"]:
            if day in sunday_cols:
                col = sunday_cols[day]
                sunday_sheet.cell(row_idx, col).value = "X"
            
            # Tìm note tương ứng với ngày này
            for note in work_info["notes"]:
                # Parse note: "2(TP HCM)" hoặc "9-16(Cần Thơ)" hoặc "2+9(An Giang)"
                match = re.match(r'([\d\-\+]+)\((.*?)\)', note)
                if match:
                    days_part = match.group(1)
                    location = match.group(2)
                    
                    # Parse ngày từ days_part
                    if '-' in days_part and '+' not in days_part:
                        day_nums = [int(x) for x in days_part.split('-')]
                        if len(day_nums) == 2:
                            note_days = list(range(day_nums[0], day_nums[1] + 1))
                        else:
                            note_days = day_nums
                    elif '+' in days_part:
                        note_days = [int(x) for x in days_part.split('+')]
                    else:
                        note_days = [int(days_part)]
                    
                    if day in note_days and f"{day}({location})" not in sunday_notes:
                        sunday_notes.append(f"{day}({location})")
        
        # Cột I: Ghi chú
        if sunday_notes:
            sunday_sheet.cell(row_idx, note_col).value = "; ".join(sunday_notes)
        
        row_idx += 1


def fill_template(template_path: Path, work_data: dict, target_month: int, target_year: int, output_path: Path):
    """
    Fill dữ liệu vào template và lưu ra output_path.
    GIỮ NGUYÊN toàn bộ structure/format của template.
    """
    wb = load_workbook(template_path)
    
    # Tìm sheet phù hợp: ưu tiên sheet có pattern "XX YYYY" (không phải "chủ nhật")
    ws = None
    target_sheet_name = f"{target_month:02d} {target_year}"
    
    # 1. Tìm sheet đúng tên tháng/năm mục tiêu
    if target_sheet_name in wb.sheetnames:
        ws = wb[target_sheet_name]
    else:
        # 2. Tìm sheet có pattern "XX YYYY" (bất kỳ tháng nào, dùng làm template)
        import re
        for sheet_name in wb.sheetnames:
            if re.match(r'\d{1,2}\s+\d{4}', sheet_name):
                ws = wb[sheet_name]
                # Đổi tên sheet theo tháng/năm mục tiêu
                ws.title = target_sheet_name
                break
        
        # 3. Nếu vẫn không tìm thấy, dùng active sheet
        if not ws:
            ws = wb.active
            ws.title = target_sheet_name
    
    # Tìm hàng header (chứa "HO VA TEN" hoặc "STT")
    header_row = None
    for row_idx in range(1, 20):
        cell_val = ws.cell(row_idx, 2).value  # Column B
        if cell_val and ("HO VA TEN" in normalize_text(str(cell_val)) or 
                        "HỌ VÀ TÊN" in normalize_text(str(cell_val))):
            header_row = row_idx
            break
    
    if not header_row:
        header_row = 6  # Default
    
    # Tìm cột bắt đầu của ngày (tìm số 1 trong header_row)
    day_start_col = None
    for col_idx in range(1, 50):
        val = ws.cell(header_row, col_idx).value
        try:
            if int(val) == 1:
                day_start_col = col_idx
                break
        except (TypeError, ValueError):
            continue
    
    if not day_start_col:
        day_start_col = 4  # Default: Column D
    
    days_in_month = calendar.monthrange(target_year, target_month)[1]
    
    # Tìm cột "Ghi chú công tác" (không phải "Ghi chú nghỉ bù")
    note_col = None
    for col_idx in range(day_start_col, 60):
        # Tìm trong cả header_row và weekday_row
        for check_row in [header_row, header_row + 1, header_row - 1]:
            val = ws.cell(check_row, col_idx).value
            if val:
                norm_val = normalize_text(str(val))
                # Phải có "CONG TAC" và "GHI CHU", nhưng KHÔNG có "NGHI BU"
                if "GHI CHU" in norm_val and "CONG TAC" in norm_val and "NGHI" not in norm_val:
                    note_col = col_idx
                    break
        if note_col:
            break
    
    if not note_col:
        note_col = day_start_col + days_in_month + 1
    
    # Tìm weekday row (row ngay sau header, có thể chứa T2, T3, CN...)
    weekday_row = header_row + 1
    
    # KHÔNG UPDATE header row nữa - giữ nguyên template
    # Chỉ update weekday nếu cell đang trống hoặc có format weekday
    weekday_labels = {0: "T2", 1: "T3", 2: "T4", 3: "T5", 4: "T6", 5: "T7", 6: "CN"}
    for day in range(1, days_in_month + 1):
        col = day_start_col + day - 1
        weekday = calendar.weekday(target_year, target_month, day)
        cell = ws.cell(weekday_row, col)
        # Chỉ update nếu cell trống hoặc đã có weekday label
        if not cell.value or normalize_text(str(cell.value)) in ["T2", "T3", "T4", "T5", "T6", "T7", "CN"]:
            cell.value = weekday_labels[weekday]
    
    # Fill màu chủ nhật
    sunday_fill = PatternFill(fill_type="solid", start_color="F5E1F3", end_color="F5E1F3")
    ct_font = Font(color="FF0000")
    
    # Iterate data rows
    empty_count = 0
    for row_idx in range(weekday_row + 1, ws.max_row + 1):
        stt_cell = ws.cell(row_idx, 1)  # Column A (STT)
        name_cell = ws.cell(row_idx, 2)  # Column B (Họ và tên)
        driver_name = name_cell.value
        
        # Dừng khi gặp 3 hàng trống liên tiếp
        if not driver_name:
            empty_count += 1
            if empty_count >= 3:
                break
            continue
        else:
            empty_count = 0
        
        driver_key = normalize_text(str(driver_name))
        
        # Dừng khi gặp các dòng ghi chú/ký hiệu
        stop_keywords = ["LUU Y", "KY HIEU", "GHI CHU", "STT", "LAM Y", "NGHI BU", "NB", "HL", 
                        "NGAY NGHI", "NGHI PHEP", "CHAM CONG", "1/2PS", "1/2PC", "DI LAM DU"]
        if any(kw in driver_key for kw in stop_keywords):
            break
        
        # Chỉ xử lý nếu cột STT là số (tài xế thật)
        try:
            int(str(stt_cell.value))
        except (ValueError, TypeError):
            # Không phải số → không phải hàng tài xế
            continue
        
        work_info = work_data.get(driver_key)
        
        # Fill các ngày
        for day in range(1, days_in_month + 1):
            col = day_start_col + day - 1
            cell = ws.cell(row_idx, col)
            
            weekday = calendar.weekday(target_year, target_month, day)
            
            if weekday == 6:  # Chủ nhật
                cell.value = None
                cell.fill = sunday_fill
            elif work_info and day in work_info["days"]:
                cell.value = "CT"
                cell.font = ct_font
            else:
                cell.value = 1
        
        # Fill ghi chú
        if work_info and work_info["notes"]:
            ws.cell(row_idx, note_col).value = "; ".join(work_info["notes"])
    
    # Update tiêu đề tháng
    # Tìm ô chứa "THÁNG" và update
    for row_idx in range(1, 10):
        for col_idx in range(1, 30):
            cell = ws.cell(row_idx, col_idx)
            if cell.value and "THÁNG" in str(cell.value).upper():
                # Update chữ "THÁNG XX" thành tháng mục tiêu
                original_text = str(cell.value)
                # Replace "THÁNG XX" hoặc "THÁNG X" bằng tháng mục tiêu
                import re
                new_text = re.sub(r'THÁNG\s+\d{1,2}', f'THÁNG {target_month}', original_text, flags=re.IGNORECASE)
                cell.value = new_text
    
    # Update ô Y3 (nếu có)
    try:
        ws["Y3"].value = f"{target_month:02d}/{target_year}"
    except:
        pass
    
    # Set active sheet là sheet tháng vừa tạo
    wb.active = wb.sheetnames.index(ws.title)
    
    # Save
    wb.save(output_path)


class ExcelApp:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("Tool xử lý Excel công tác")
        self.root.geometry("600x400")
        self.root.configure(bg="#f0f5f0")
        
        try:
            logo_path = Path(__file__).parent / "logo.png"
            ico_path = Path(__file__).parent / "logo.ico"
            
            if logo_path.exists() and Image:
                if not ico_path.exists():
                    img = Image.open(logo_path)
                    img = img.resize((64, 64), Image.Resampling.LANCZOS)
                    img.save(ico_path, format='ICO', sizes=[(64, 64), (32, 32), (16, 16)])
                
                self.root.iconbitmap(str(ico_path))
            elif ico_path.exists():
                self.root.iconbitmap(str(ico_path))
        except Exception:
            pass
        
        self.input_path: Optional[Path] = None
        
        self.status_var = tk.StringVar(value="Chưa chọn file.")
        
        title_label = tk.Label(root, text="TOOL XỬ LÝ FILE CÔNG TÁC", 
                              font=("Arial", 18, "bold"), 
                              bg="#00784a", fg="white", 
                              pady=15)
        title_label.pack(fill="x", padx=0, pady=0)
        
        btn_select = tk.Button(root, text="📂 Thêm file excel", width=30, 
                              font=("Arial", 12, "bold"),
                              bg="#4fb258", fg="white", 
                              activebackground="#3d9146",
                              relief="flat", cursor="hand2",
                              pady=10,
                              command=self.select_input)
        btn_select.pack(padx=20, pady=20)
        
        frame_date = tk.Frame(root, bg="#f0f5f0")
        frame_date.pack(padx=20, pady=15)
        
        tk.Label(frame_date, text="Tháng:", bg="#f0f5f0", 
                font=("Arial", 12)).grid(row=0, column=0, padx=8)
        self.month_var = tk.IntVar(value=datetime.now().month)
        self.month_spinbox = tk.Spinbox(frame_date, from_=1, to=12, 
                                       textvariable=self.month_var, width=12,
                                       font=("Arial", 12))
        self.month_spinbox.grid(row=0, column=1, padx=8)
        
        tk.Label(frame_date, text="Năm:", bg="#f0f5f0", 
                font=("Arial", 12)).grid(row=0, column=2, padx=8)
        self.year_var = tk.IntVar(value=datetime.now().year)
        self.year_spinbox = tk.Spinbox(frame_date, from_=2020, to=2099, 
                                      textvariable=self.year_var, width=12,
                                      font=("Arial", 12))
        self.year_spinbox.grid(row=0, column=3, padx=8)
        
        btn_export = tk.Button(root, text="✅ Xuất file excel", width=30, 
                              font=("Arial", 12, "bold"),
                              bg="#00784a", fg="white", 
                              activebackground="#005a38",
                              relief="flat", cursor="hand2",
                              pady=10,
                              command=self.export_file)
        btn_export.pack(padx=20, pady=20)
        
        status_frame = tk.Frame(root, bg="white", relief="solid", borderwidth=1)
        status_frame.pack(fill="x", padx=20, pady=15)
        
        lbl_status = tk.Label(status_frame, textvariable=self.status_var, 
                             wraplength=560, justify="left",
                             bg="white", fg="#333333",
                             font=("Arial", 10), pady=12, padx=12)
        lbl_status.pack(fill="x")
    
    def select_input(self):
        path = filedialog.askopenfilename(
            title="Chọn file Lệnh điều động",
            filetypes=[("Excel files", "*.xlsx *.xlsm"), ("All files", "*.*")]
        )
        if not path:
            return
        
        self.input_path = Path(path)
        
        month_year = parse_month_from_filename(self.input_path)
        if month_year:
            month, year = month_year
            self.month_var.set(month)
            self.year_var.set(year)
            self.status_var.set(f"Đã chọn: {self.input_path.name}\nTháng: {month}/{year}")
        else:
            self.status_var.set(f"Đã chọn: {self.input_path.name}")
    
    def export_file(self):
        if load_workbook is None:
            messagebox.showerror("Thiếu thư viện", "Vui lòng cài openpyxl:\npip install openpyxl")
            return
        
        if not self.input_path:
            messagebox.showwarning("Thiếu file", "Vui lòng chọn file Lệnh điều động trước.")
            return
        
        target_month = self.month_var.get()
        target_year = self.year_var.get()
        
        try:
            # Đọc dữ liệu công tác
            work_data = collect_work_data(self.input_path, target_month, target_year)
            
            if not work_data:
                messagebox.showwarning("Không có dữ liệu", f"Không tìm thấy dữ liệu công tác tháng {target_month}/{target_year}")
                return
            
            # Tìm template
            template_path = find_template_file(self.input_path.parent, target_month, target_year)
            if not template_path:
                messagebox.showerror("Thiếu file mẫu", 
                    f"Không tìm thấy file mẫu 'Cham cong {target_month:02d} {target_year}.xlsx'")
                return
            
            # Tạo output path
            output_path = self.input_path.parent / f"Cham cong {target_month:02d}-{target_year}_ket_qua.xlsx"
            counter = 1
            while output_path.exists():
                output_path = self.input_path.parent / f"Cham cong {target_month:02d}-{target_year}_ket_qua_v{counter}.xlsx"
                counter += 1
            
            # Fill template
            fill_template(template_path, work_data, target_month, target_year, output_path)
            
            # Fill sheet "Chủ Nhật"
            wb = load_workbook(output_path)
            fill_sunday_sheet(wb, work_data, target_month, target_year)
            wb.save(output_path)
            wb.close()
            
            self.status_var.set(f"Xuất thành công!\n{output_path.name}")
            messagebox.showinfo("Hoàn tất", f"Đã lưu file:\n{output_path.name}")
            
        except Exception as e:
            messagebox.showerror("Lỗi", f"Có lỗi xảy ra:\n{str(e)}")
            import traceback
            traceback.print_exc()


def main():
    root = tk.Tk()
    app = ExcelApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
